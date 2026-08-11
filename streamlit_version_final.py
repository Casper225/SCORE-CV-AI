# -*- coding: utf-8 -*-
"""Score CV AI — Dashboard professionnel d'analyse de compatibilité CV / offre."""

from __future__ import annotations

import base64
import html
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List

import streamlit as st

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    from docx import Document
except ImportError:
    Document = None

try:
    from PIL import Image
except ImportError:
    Image = None

try:
    import easyocr
except ImportError:
    easyocr = None

try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
except ImportError:
    TfidfVectorizer = None
    cosine_similarity = None

try:
    import requests
except ImportError:
    requests = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
    Alignment = Font = PatternFill = None
    get_column_letter = None


# -----------------------------------------------------------------------------
# Configuration
# -----------------------------------------------------------------------------

st.set_page_config(
    page_title="Score CV AI | Analyse intelligente",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="expanded",
)

PRIMARY = "#635BFF"
DARK = "#101828"
MUTED = "#667085"
SURFACE = "#FFFFFF"
CANVAS = "#F6F8FC"
BORDER = "#E4E7EC"
SUCCESS = "#079455"
WARNING = "#DC6803"
DANGER = "#D92D20"
APP_DIR = Path(__file__).resolve().parent
BANNER_PATH = APP_DIR / "assets" / "scorecv_banner.png"


# -----------------------------------------------------------------------------
# Design system
# -----------------------------------------------------------------------------

st.markdown(
    f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=Space+Grotesk:wght@500;600;700&display=swap');

    :root {{
        --primary: {PRIMARY};
        --dark: {DARK};
        --muted: {MUTED};
        --surface: {SURFACE};
        --canvas: {CANVAS};
        --border: {BORDER};
    }}

    html, body, [class*="css"] {{
        font-family: 'DM Sans', sans-serif;
    }}

    .stApp {{
        background: var(--canvas);
        color: var(--dark);
    }}

    .main .block-container {{
        max-width: 1480px;
        padding: 2.2rem 3rem 1.5rem;
    }}

    [data-testid="stSidebar"] {{
        background: #111827;
        border-right: 1px solid #1F2937;
    }}

    [data-testid="stSidebar"] * {{ color: #F9FAFB !important; }}
    [data-testid="stSidebar"] hr {{ border-color: #374151; }}
    [data-testid="stSidebar"] .stCaption {{ color: #CBD5E1 !important; }}

    h1, h2, h3, h4 {{
        font-family: 'Space Grotesk', sans-serif;
        color: var(--dark);
        letter-spacing: -0.03em;
    }}

    .hero-wrap {{
        position: relative;
        aspect-ratio: 1553 / 688;
        min-height: 0;
        margin-bottom: 1.15rem;
        overflow: hidden;
        border-radius: 22px;
        border: 1px solid rgba(255,255,255,.42);
        background: #08243A;
        box-shadow: 0 18px 42px rgba(16,24,40,.12);
    }}
    .hero-wrap img {{
        display: block;
        width: 100%;
        height: 100%;
        object-fit: contain;
        object-position: center;
    }}
    .hero-overlay {{
        position: absolute;
        inset: 0;
        display: flex;
        align-items: flex-end;
        padding: 1.8rem 2rem;
        background: linear-gradient(90deg, rgba(3,18,33,.86) 0%, rgba(3,18,33,.32) 55%, rgba(3,18,33,.08) 100%);
    }}
    .hero-copy {{ max-width: 560px; color: #fff; }}
    .hero-kicker {{ margin-bottom: .55rem; color: #A7F3D0; font-size: .72rem; font-weight: 700; letter-spacing: .16em; text-transform: uppercase; }}
    .hero-title {{ margin: 0; color: #fff; font-family: 'Space Grotesk'; font-size: clamp(1.8rem, 3vw, 3rem); line-height: 1.02; font-weight: 700; letter-spacing: -.045em; }}
    .hero-description {{ margin: .7rem 0 0; color: #E0F2FE; font-size: .98rem; line-height: 1.45; }}

    .brand-bar {{
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 24px;
        padding: 1.35rem 1.55rem;
        margin-bottom: 1.15rem;
        background: linear-gradient(135deg, #FFFFFF 0%, #F7F7FF 100%);
        border: 1px solid var(--border);
        border-radius: 18px;
    }}

    .brand-left {{ display: flex; align-items: center; gap: 14px; }}
    .brand-mark {{
        width: 46px; height: 46px; display: grid; place-items: center;
        border-radius: 14px; color: #fff; font-weight: 800; font-size: 21px;
        background: linear-gradient(135deg, #635BFF, #8B5CF6);
        box-shadow: 0 9px 22px rgba(99,91,255,.24);
    }}
    .brand-title {{ margin: 0; font-family: 'Space Grotesk'; font-size: 1.55rem; font-weight: 700; color: var(--dark); }}
    .brand-subtitle {{ margin: 3px 0 0; color: var(--muted); font-size: .88rem; }}
    .brand-status {{
        padding: 7px 12px; border-radius: 999px; font-size: .75rem; font-weight: 700;
        color: #4338CA; background: #EEF2FF; border: 1px solid #C7D2FE;
        white-space: nowrap;
    }}

    .section-label {{
        margin: 1.3rem 0 .65rem; color: var(--primary); font-size: .72rem;
        font-weight: 700; letter-spacing: .12em; text-transform: uppercase;
    }}

    .panel {{
        min-height: 0; padding: 1.1rem 1.2rem; background: var(--surface);
        border: 1px solid var(--border); border-radius: 16px;
        box-shadow: 0 5px 18px rgba(16,24,40,.035);
    }}
    [data-testid="stVerticalBlockBorderWrapper"] {{
        background: var(--surface);
        border: 1px solid var(--border);
        border-radius: 16px;
        box-shadow: 0 5px 18px rgba(16,24,40,.035);
        padding: .95rem 1.05rem .8rem;
    }}
    [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stFileUploader"] {{
        background: #FAFAFF;
    }}
    /* Grille d’entrée : les deux panneaux conservent la même hauteur */
    [data-testid="stHorizontalBlock"]:has([data-testid="stFileUploader"]) [data-testid="stVerticalBlockBorderWrapper"] {{
        min-height: 220px;
        box-sizing: border-box;
    }}
    [data-testid="stHorizontalBlock"]:has([data-testid="stFileUploader"]) [data-testid="stTextArea"] textarea {{
        min-height: 118px;
    }}
    .overview-panel {{ min-height: 236px; box-sizing: border-box; }}
    .result-panel {{ min-height: 156px; box-sizing: border-box; }}
    [data-testid="stDataFrame"] {{
        border: 1px solid var(--border);
        border-radius: 12px;
        overflow: hidden;
        background: var(--surface);
    }}

    .panel-title {{ margin: 0 0 .85rem; font-family: 'Space Grotesk'; font-size: 1rem; font-weight: 700; color: var(--dark); }}
    .panel-kicker {{ margin: 0 0 .25rem; font-size: .69rem; font-weight: 700; letter-spacing: .08em; color: var(--muted); text-transform: uppercase; }}

    .score-panel {{
        min-height: 236px; padding: 1.3rem; color: #fff; border: 0; border-radius: 18px;
        background: linear-gradient(145deg, #312E81 0%, #635BFF 55%, #8B5CF6 100%);
        box-shadow: 0 14px 32px rgba(99,91,255,.22);
    }}
    .score-panel .panel-kicker {{ color: #DDE1FF; }}
    .score-number {{ margin: .2rem 0 .45rem; font-family: 'Space Grotesk'; font-size: 3.75rem; line-height: 1; font-weight: 700; }}
    .score-status {{ display: inline-block; padding: 5px 10px; border-radius: 999px; background: rgba(255,255,255,.15); font-size: .76rem; font-weight: 700; }}
    .score-progress {{ height: 7px; margin-top: 1.3rem; border-radius: 99px; background: rgba(255,255,255,.22); overflow: hidden; }}
    .score-progress > div {{ height: 100%; border-radius: inherit; background: #FFFFFF; }}

    .metric-box {{ padding: 1rem 1.1rem; background: var(--surface); border: 1px solid var(--border); border-radius: 16px; min-height: 0; }}
    .metric-label {{ color: var(--muted); font-size: .76rem; font-weight: 600; }}
    .metric-value {{ margin-top: .4rem; color: var(--dark); font-family: 'Space Grotesk'; font-size: 1.65rem; font-weight: 700; }}
    .metric-caption {{ color: var(--muted); font-size: .75rem; }}

    .info-row {{ display:flex; justify-content:space-between; gap:12px; padding:.62rem 0; border-bottom:1px solid #F2F4F7; font-size:.83rem; }}
    .info-row:last-child {{ border-bottom:0; }}
    .info-label {{ color:var(--muted); }}
    .info-value {{ color:var(--dark); font-weight:600; text-align:right; overflow-wrap:anywhere; }}

    .skill-wrap {{ display:flex; flex-wrap:wrap; gap:8px; }}
    .skill {{ display:inline-flex; align-items:center; padding:7px 10px; border-radius:9px; font-size:.78rem; font-weight:600; border:1px solid; }}
    .skill.ok {{ color:#067647; background:#ECFDF3; border-color:#ABEFC6; }}
    .skill.missing {{ color:#B42318; background:#FEF3F2; border-color:#FECDCA; }}
    .skill.neutral {{ color:#344054; background:#F8FAFC; border-color:#E4E7EC; }}

    .empty-state {{ padding: 3rem 1rem; text-align:center; color:var(--muted); background:#FFFFFF; border:1px dashed #CBD5E1; border-radius:16px; }}
    .empty-state strong {{ display:block; margin-bottom:.35rem; color:var(--dark); font-family:'Space Grotesk'; font-size:1.05rem; }}

    .footer {{
        margin-top: 2rem; padding: 1.25rem 0 .25rem; border-top: 1px solid var(--border);
        display:flex; align-items:center; justify-content:space-between; gap:16px; color:var(--muted); font-size:.78rem;
    }}
    .footer a {{ color:var(--primary); text-decoration:none; font-weight:600; }}
    .footer a:hover {{ text-decoration:underline; }}

    .stButton > button {{
        min-height: 46px; border:0; border-radius:12px; color:#fff; background:var(--primary);
        font-weight:700; box-shadow:0 8px 18px rgba(99,91,255,.2);
    }}
    .stButton > button:hover {{ background:#5148E8; color:#fff; border:0; }}
    .stDownloadButton > button {{ border-radius:10px; }}
    .stTextArea textarea, .stTextInput input {{ border-radius:10px; border-color:#D0D5DD; }}
    [data-testid="stFileUploader"] {{ background:#FAFAFF; border:1px dashed #A5B4FC; border-radius:12px; padding:.35rem; }}
    </style>
    """,
    unsafe_allow_html=True,
)


# -----------------------------------------------------------------------------
# Taxonomies and utilities
# -----------------------------------------------------------------------------

SKILL_TAXONOMY = [
    "Python", "Docker", "Kubernetes", "Git", "Data Science", "NLP", "Spark",
    "Cloud", "SQL", "Machine Learning", "Deep Learning", "TensorFlow", "PyTorch",
    "Scikit-learn", "Pandas", "NumPy", "Power BI", "Excel", "AWS", "Azure",
    "GCP", "Java", "React",
]

SOFT_SKILLS_TAXONOMY = [
    "leadership", "communication", "teamwork", "organisation", "autonomie",
    "adaptabilité", "créativité", "résolution de problèmes", "travail d'équipe",
    "gestion d'équipe",
]

LANGUAGE_MAP = {
    "français": "Français", "francais": "Français", "anglais": "Anglais",
    "english": "Anglais", "espagnol": "Espagnol", "allemand": "Allemand",
}

DEFAULT_JOB = """Senior Data Scientist & Lead Engineer.
Expérience exigée : 3 ans minimum.

Compétences clés recherchées :
- Python, SQL, NLP, Data Science
- Scikit-learn, Machine Learning, Deep Learning
- Docker, Git, Cloud (AWS ou GCP)
- Esprit d'équipe, autonomie et communication.
"""


def esc(value: Any) -> str:
    return html.escape(str(value))


def contains_term(text: str, term: str) -> bool:
    pattern = r"(?<!\w)" + re.escape(term.lower()) + r"(?!\w)"
    return bool(re.search(pattern, text.lower()))


def clean_text(raw_text: str) -> str:
    text = raw_text.lower()
    text = re.sub(r"http\S+|www\S+|<.*?>", " ", text)
    text = re.sub(r"[^\w\s\+\#\.-]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    stopwords = {
        "le", "la", "les", "un", "une", "des", "du", "de", "et", "en", "pour",
        "avec", "dans", "sur", "par", "est", "sont", "au", "aux", "the", "and",
        "or", "in", "at", "for", "with",
    }
    return " ".join(word for word in text.split() if word not in stopwords)


@st.cache_resource(show_spinner=False)
def get_ocr_reader():
    if easyocr is None:
        return None
    return easyocr.Reader(["fr", "en"], gpu=False, verbose=False)


def extract_text(uploaded_file) -> str:
    if uploaded_file is None:
        return ""
    filename = uploaded_file.name.lower()
    try:
        if filename.endswith(".pdf"):
            if pdfplumber is None:
                return ""
            with pdfplumber.open(uploaded_file) as pdf:
                return "\n".join((page.extract_text() or "") for page in pdf.pages).strip()
        if filename.endswith(".docx"):
            if Document is None:
                return ""
            document = Document(uploaded_file)
            return "\n".join(p.text for p in document.paragraphs if p.text.strip()).strip()
        if filename.endswith(".txt"):
            return uploaded_file.getvalue().decode("utf-8", errors="ignore").strip()
        if filename.endswith((".png", ".jpg", ".jpeg")):
            reader = get_ocr_reader()
            if reader is None or Image is None:
                return ""
            image = Image.open(BytesIO(uploaded_file.getvalue()))
            results = reader.readtext(image, detail=0, paragraph=True)
            return "\n".join(results).strip()
    except Exception as error:
        st.error(f"Erreur d'extraction : {error}")
    return ""


def extract_entities(raw_text: str, cleaned_text: str) -> Dict[str, Any]:
    entities: Dict[str, Any] = {
        "candidate_name": "Candidat non identifié", "email": "Non détecté",
        "phone": "Non détecté", "years_experience": 0, "skills": [],
        "soft_skills": [], "languages": [],
    }
    if not raw_text:
        return entities

    email_match = re.search(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", raw_text)
    if email_match:
        entities["email"] = email_match.group(0)

    phone_match = re.search(r"(?:\+?\d[\d\s().-]{7,}\d)", raw_text)
    if phone_match:
        entities["phone"] = phone_match.group(0).strip()

    experience_values = re.findall(
        r"(\d+)\s*(?:ans?|years?)\s*(?:d['’]?expérience|of experience)?",
        raw_text,
        re.IGNORECASE,
    )
    valid_values = [int(value) for value in experience_values if int(value) < 40]
    if valid_values:
        entities["years_experience"] = max(valid_values)

    lines = [re.sub(r"^[•|\-–—]+\s*", "", line.strip()) for line in raw_text.splitlines() if line.strip()]
    name_labels = re.compile(r"^(?:nom(?: complet)?|name|candidat|candidate)\s*[:\-–—]\s*(.+)$", re.IGNORECASE)
    excluded = {
        "curriculum vitae", "resume", "cv", "profil professionnel", "professional profile",
        "coordonnées", "contact", "expérience professionnelle", "formation", "compétences",
        "skills", "education", "expérience", "experience",
    }
    for line in lines[:15]:
        labeled = name_labels.match(line)
        if labeled:
            candidate = labeled.group(1).strip()
            if candidate:
                entities["candidate_name"] = candidate.title()
                break
        normalized = re.sub(r"[^a-zA-ZÀ-ÿ' -]", "", line).strip()
        words = normalized.split()
        lowered = normalized.lower()
        if (
            2 <= len(words) <= 4
            and lowered not in excluded
            and not any(char.isdigit() for char in line)
            and "@" not in line
            and not any(token in lowered for token in ("téléphone", "phone", "linkedin", "github", "http"))
            and len(normalized) >= 5
        ):
            entities["candidate_name"] = " ".join(word.capitalize() for word in words)
            break

    entities["skills"] = [skill for skill in SKILL_TAXONOMY if contains_term(cleaned_text, skill)]
    entities["soft_skills"] = [skill.capitalize() for skill in SOFT_SKILLS_TAXONOMY if contains_term(cleaned_text, skill)]
    entities["languages"] = list(dict.fromkeys(language for key, language in LANGUAGE_MAP.items() if contains_term(cleaned_text, key)))
    return entities


def extract_relationships(entities: Dict[str, Any], job_cleaned: str) -> Dict[str, Any]:
    required = [skill for skill in SKILL_TAXONOMY if contains_term(job_cleaned, skill)]
    matched = [{"skill": skill} for skill in required if skill in entities["skills"]]
    missing = [{"skill": skill} for skill in required if skill not in entities["skills"]]
    return {"required_skills_in_job": required, "matched_relationships": matched, "missing_relationships": missing}


def compute_similarity(cv_clean: str, job_clean: str) -> float:
    if TfidfVectorizer is None or cosine_similarity is None or not cv_clean or not job_clean:
        return 0.0
    try:
        vectorizer = TfidfVectorizer(ngram_range=(1, 2), min_df=1, sublinear_tf=True)
        matrix = vectorizer.fit_transform([cv_clean, job_clean])
        return float(cosine_similarity(matrix[0:1], matrix[1:2])[0][0] * 100)
    except Exception:
        return 0.0


def compute_score(entities: Dict[str, Any], relationships: Dict[str, Any], similarity: float) -> tuple[int, Dict[str, float]]:
    total_required = len(relationships["required_skills_in_job"])
    matched_required = len(relationships["matched_relationships"])
    skill_ratio = matched_required / total_required if total_required else 0.0
    soft_score = min(len(entities["soft_skills"]) / 3.0, 1.0)
    score = (similarity * .45) + (skill_ratio * 100 * .45) + (soft_score * 100 * .10)
    return int(round(max(0, min(100, score)))), {"skill_match_ratio": skill_ratio, "soft_skills_score": soft_score}


def search_market(job_title: str, api_key: str) -> Dict[str, Any]:
    if not api_key or requests is None:
        return {"status": "NOT_CONFIGURED", "market_demand": "Inconnue", "results_found": 0}
    try:
        response = requests.get(
            "https://jsearch.p.rapidapi.com/search",
            headers={"X-RapidAPI-Key": api_key, "X-RapidAPI-Host": "jsearch.p.rapidapi.com"},
            params={"query": job_title, "page": "1", "num_pages": "1"},
            timeout=8,
        )
        response.raise_for_status()
        count = len(response.json().get("data", []))
        return {"status": "LIVE", "market_demand": "Très élevée" if count >= 8 else "Modérée", "results_found": count}
    except Exception:
        return {"status": "ERROR", "market_demand": "Indisponible", "results_found": 0}


def status_for_score(score: int) -> tuple[str, str]:
    if score >= 75:
        return "Excellent match", SUCCESS
    if score >= 50:
        return "Profil intermédiaire", WARNING
    return "Compatibilité faible", DANGER


def read_secret(name: str, default: str = "") -> str:
    """Lit un secret sans faire planter l'application si secrets.toml est absent."""
    try:
        return str(st.secrets.get(name, default))
    except Exception:
        return default


def info_row(label: str, value: Any) -> str:
    return f'<div class="info-row"><span class="info-label">{esc(label)}</span><span class="info-value">{esc(value)}</span></div>'


def build_excel_report(result: Dict[str, Any]) -> bytes:
    """Construit un rapport Excel lisible et prêt à partager."""
    if Workbook is None:
        raise RuntimeError("La dépendance openpyxl est absente. Ajoutez openpyxl à requirements.txt.")

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Résumé"
    details = workbook.create_sheet("Compétences")

    navy = "102A43"
    green = "2E8B68"
    light_blue = "EAF2F8"
    light_green = "EAF7F0"
    gray = "667085"

    summary.merge_cells("A1:D1")
    summary["A1"] = "SCORE CV AI — RAPPORT D'ANALYSE"
    summary["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor=navy)
    summary["A1"].alignment = Alignment(horizontal="center")

    result_entities = result["entities"]
    summary_rows = [
        ("Score global", f'{result["score"]}%'),
        ("Statut", result.get("status_label", "")),
        ("Fichier analysé", result["filename"]),
        ("Candidat", result_entities["candidate_name"]),
        ("Email", result_entities["email"]),
        ("Téléphone", result_entities["phone"]),
        ("Expérience", f'{result_entities["years_experience"]} an(s)'),
        ("Similarité textuelle", f'{result["tfidf"]:.1f}%'),
        ("Compétences couvertes", f'{result["features"]["skill_match_ratio"] * 100:.0f}%'),
        ("Soft skills", f'{result["features"]["soft_skills_score"] * 100:.0f}%'),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=3):
        summary.cell(row_index, 1, label)
        summary.cell(row_index, 2, value)
        summary.cell(row_index, 1).font = Font(bold=True, color=gray)
        summary.cell(row_index, 2).font = Font(bold=row_index == 3, color=navy if row_index == 3 else "1D2939")
        if row_index == 3:
            summary.cell(row_index, 2).fill = PatternFill("solid", fgColor=light_green)

    details.append(["Compétence", "Statut"])
    for cell in details[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=green)
    matched = {item["skill"] for item in result["relationships"]["matched_relationships"]}
    missing = {item["skill"] for item in result["relationships"]["missing_relationships"]}
    for skill in sorted(matched | missing):
        details.append([skill, "Confirmée" if skill in matched else "À renforcer"])
    for row in details.iter_rows(min_row=2, max_col=2):
        row[1].fill = PatternFill("solid", fgColor=light_green if row[1].value == "Confirmée" else "FCEBEA")

    for sheet in (summary, details):
        sheet.freeze_panes = "A3" if sheet.title == "Résumé" else "A2"
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 42
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def skill_html(items: List[str], kind: str) -> str:
    if not items:
        return '<span style="color:#98A2B3;font-size:.82rem">Aucun élément détecté.</span>'
    return '<div class="skill-wrap">' + "".join(f'<span class="skill {kind}">{esc(item)}</span>' for item in items) + "</div>"


# -----------------------------------------------------------------------------
# Session state
# -----------------------------------------------------------------------------

st.session_state.setdefault("history", [])
st.session_state.setdefault("show_history", False)


# -----------------------------------------------------------------------------
# Sidebar
# -----------------------------------------------------------------------------

with st.sidebar:
    st.markdown("## Score CV AI")
    st.caption("Analyse intelligente du matching candidat–poste")
    st.markdown("---")
    if st.button("Historique des analyses", use_container_width=True):
        st.session_state["show_history"] = not st.session_state["show_history"]
    if st.session_state["history"]:
        st.caption(f'{len(st.session_state["history"])} analyse(s) enregistrée(s) dans cette session')
    st.markdown("---")
    st.markdown("### Parcours recommandé")
    st.caption("1. Importez un CV.\n2. Collez l'offre d'emploi.\n3. Lancez l'analyse.\n4. Explorez le score et les écarts.")
    st.markdown("---")
    st.caption("Formats pris en charge : PDF, DOCX, TXT, PNG, JPG")


# -----------------------------------------------------------------------------
# Header and inputs
# -----------------------------------------------------------------------------

if BANNER_PATH.exists():
    banner_b64 = base64.b64encode(BANNER_PATH.read_bytes()).decode("ascii")
    st.markdown(
        f"""
        <div class="hero-wrap">
          <img src="data:image/png;base64,{banner_b64}" alt="Bannière ScoreCV" />
        </div>
        """,
        unsafe_allow_html=True,
    )
else:
    st.warning("Bannière introuvable : placez scorecv_banner.png dans le dossier assets.")

st.markdown(
    """
    <div class="brand-bar">
      <div class="brand-left">
        <div class="brand-mark">◈</div>
        <div>
          <h1 class="brand-title">Score CV AI</h1>
          <p class="brand-subtitle">Plateforme d'analyse intelligente du matching candidat–poste</p>
        </div>
      </div>
      <div class="brand-status">NLP · Scoring · Matching</div>
    </div>
    """,
    unsafe_allow_html=True,
)

st.markdown('<div class="section-label">Nouvelle analyse</div>', unsafe_allow_html=True)
input_cv, input_job = st.columns([0.9, 1.1], gap="large")

with input_cv:
    with st.container(border=True):
        st.markdown('<div class="panel-title">Document candidat</div>', unsafe_allow_html=True)
        st.caption("Importez le CV à analyser")
        uploaded_file = st.file_uploader(
            "CV",
            type=["pdf", "docx", "txt", "png", "jpg", "jpeg"],
            label_visibility="collapsed",
        )

with input_job:
    with st.container(border=True):
        st.markdown("<div class='panel-title'>Offre d'emploi</div>", unsafe_allow_html=True)
        st.caption("Décrivez le poste et les compétences recherchées")
        job_text = st.text_area("Description", value="", height=118, placeholder="Collez ici la description du poste et les compétences recherchées…", label_visibility="collapsed")

st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)
run_pipeline = st.button("Lancer l'analyse du profil", use_container_width=True)


# -----------------------------------------------------------------------------
# Pipeline
# -----------------------------------------------------------------------------

if run_pipeline:
    if uploaded_file is None:
        st.warning("Importez un CV avant de lancer l'analyse.")
        st.stop()
    if not job_text.strip():
        st.warning("Ajoutez une description de poste avant de lancer l'analyse.")
        st.stop()

    with st.spinner("Extraction, NLP et scoring en cours…"):
        cv_raw = extract_text(uploaded_file)
        if not cv_raw:
            st.error("Le contenu du CV n'a pas pu être extrait. Vérifiez le format ou la qualité du document.")
            st.stop()

        cv_clean = clean_text(cv_raw)
        job_clean = clean_text(job_text)
        entities = extract_entities(cv_raw, cv_clean)
        relationships = extract_relationships(entities, job_clean)
        similarity = compute_similarity(cv_clean, job_clean)
        score, features = compute_score(entities, relationships, similarity)
        analysis_result = {
            "score": score,
            "tfidf": similarity,
            "entities": entities,
            "relationships": relationships,
            "features": features,
            "filename": uploaded_file.name,
        }
        st.session_state["results"] = analysis_result
        st.session_state["history"].insert(0, {
            "date": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "filename": uploaded_file.name,
            "candidate_name": entities.get("candidate_name", "Candidat non identifié"),
            "score": score,
            "status": status_for_score(score)[0],
        })
        st.session_state["history"] = st.session_state["history"][:25]


# -----------------------------------------------------------------------------
# History dashboard
# -----------------------------------------------------------------------------

if st.session_state["show_history"]:
    st.markdown('<div class="section-label">Historique des analyses</div>', unsafe_allow_html=True)
    if st.session_state["history"]:
        history_rows = [
            {
                "Date": item["date"],
                "CV analysé": item["filename"],
                "Candidat": item["candidate_name"],
                "Score": f'{item["score"]}%',
                "Statut": item["status"],
            }
            for item in st.session_state["history"]
        ]
        table_height = min(360, max(120, 38 * len(history_rows) + 42))
        st.dataframe(history_rows, use_container_width=True, hide_index=True, height=table_height)
    else:
        st.info("Aucune analyse enregistrée pour le moment. Lancez une analyse pour alimenter l’historique.")


# -----------------------------------------------------------------------------
# Results dashboard
# -----------------------------------------------------------------------------

if "results" not in st.session_state:
    st.markdown(
        '<div class="empty-state"><strong>Votre tableau de bord apparaîtra ici</strong>Importez un CV et lancez l’analyse pour obtenir le score de compatibilité, les compétences détectées et les recommandations.</div>',
        unsafe_allow_html=True,
    )
else:
    result = st.session_state["results"]
    score = int(result["score"])
    status_label, status_color = status_for_score(score)
    result["status_label"] = status_label
    entities = result["entities"]
    relationships = result["relationships"]

    st.markdown('<div class="section-label">Vue d’ensemble</div>', unsafe_allow_html=True)
    score_col, profile_col = st.columns([0.9, 1.25], gap="large")

    with score_col:
        st.markdown(
            f"""
            <div class="score-panel">
              <div class="panel-kicker">Score global de compatibilité</div>
              <div class="score-number">{score}%</div>
              <div class="score-status">{esc(status_label)}</div>
              <div class="score-progress"><div style="width:{score}%"></div></div>
              <div style="display:flex;justify-content:space-between;margin-top:.55rem;color:#E0E7FF;font-size:.74rem"><span>0</span><span>100</span></div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with profile_col:
        profile_html = (
            '<div class="panel overview-panel">'
            '<div class="panel-kicker">Profil extrait</div>'
            '<div class="panel-title">Informations détectées</div>'
            + info_row("Candidat", entities["candidate_name"])
            + info_row("Email", entities["email"])
            + info_row("Téléphone", entities["phone"])
            + info_row("Expérience", f'{entities["years_experience"]} an(s)')
            + '</div>'
        )
        st.markdown(profile_html, unsafe_allow_html=True)

    st.markdown('<div class="section-label">Indicateurs de matching</div>', unsafe_allow_html=True)
    metric_a, metric_b, metric_c, metric_d = st.columns(4, gap="medium")
    metrics = [
        ("Similarité textuelle", f'{result["tfidf"]:.1f}%', "TF-IDF + cosinus"),
        ("Compétences couvertes", f'{result["features"]["skill_match_ratio"] * 100:.0f}%', "Exigences détectées"),
        ("Soft skills", f'{result["features"]["soft_skills_score"] * 100:.0f}%', "Signaux comportementaux"),
        ("Fichier analysé", result["filename"], "Document source"),
    ]
    for column, (label, value, caption) in zip([metric_a, metric_b, metric_c, metric_d], metrics):
        with column:
            st.markdown(f'<div class="metric-box"><div class="metric-label">{esc(label)}</div><div class="metric-value">{esc(value)}</div><div class="metric-caption">{esc(caption)}</div></div>', unsafe_allow_html=True)

    st.markdown('<div class="section-label">Analyse des compétences</div>', unsafe_allow_html=True)
    skills_ok, skills_missing = st.columns(2, gap="large")
    with skills_ok:
        st.markdown(
            '<div class="panel result-panel"><div class="panel-kicker" style="color:#067647">Correspondances positives</div>'
            '<div class="panel-title">Compétences confirmées</div>'
            + skill_html([item["skill"] for item in relationships["matched_relationships"]], "ok")
            + '</div>',
            unsafe_allow_html=True,
        )
    with skills_missing:
        st.markdown(
            '<div class="panel result-panel"><div class="panel-kicker" style="color:#B42318">Écarts à traiter</div>'
            '<div class="panel-title">Compétences manquantes</div>'
            + skill_html([item["skill"] for item in relationships["missing_relationships"]], "missing")
            + '</div>',
            unsafe_allow_html=True,
        )

    st.markdown('<div class="section-label">Profil enrichi</div>', unsafe_allow_html=True)
    soft_col, lang_col, export_col = st.columns([1, 1, 1], gap="large")
    with soft_col:
        st.markdown(
            '<div class="panel result-panel"><div class="panel-kicker">Comportemental</div><div class="panel-title">Soft skills détectées</div>'
            + skill_html(entities["soft_skills"], "neutral") + '</div>',
            unsafe_allow_html=True,
        )
    with lang_col:
        st.markdown(
            '<div class="panel result-panel"><div class="panel-kicker">Communication</div><div class="panel-title">Langues détectées</div>'
            + skill_html(entities["languages"], "neutral") + '</div>',
            unsafe_allow_html=True,
        )
    with export_col:
        with st.container(border=True):
            st.markdown('<div class="panel-kicker">Traçabilité</div><div class="panel-title">Exporter le résultat</div>', unsafe_allow_html=True)
            excel_data = build_excel_report(result)
            st.download_button(
                "Télécharger le rapport Excel",
                data=excel_data,
                file_name="score_cv_ai_rapport.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )


# -----------------------------------------------------------------------------
# Footer
# -----------------------------------------------------------------------------

st.markdown(
    """
    <div class="footer">
      <span>© 2026 Score CV AI · Conçu par Lionel Kouakou</span>
      <span><a href="mailto:kouakouericlionel@gmail.com">kouakouericlionel@gmail.com</a> · <a href="https://github.com/Casper225" target="_blank">GitHub : Casper225 (Lionel KOUAKOU)</a></span>
    </div>
    """,
    unsafe_allow_html=True,
)
